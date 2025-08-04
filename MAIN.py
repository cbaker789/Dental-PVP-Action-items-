import os
import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --- Format sheet as Excel table with autosize ---
def format_sheet_as_table(sheet, df):
    n_rows, n_cols = df.shape
    last_col = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{n_rows + 1}"
    tab = Table(displayName=sheet.title.replace(" ", ""), ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max(10, min(max_length + 2, 40))

# --- Save DataFrame to Excel with formatting ---
def export_to_excel_simple(df_dict, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(output_path)
    for sheet_name in df_dict:
        sheet = wb[sheet_name]
        format_sheet_as_table(sheet, df_dict[sheet_name])
    wb.save(output_path)
    print(f"[✓] Excel saved: {output_path}")

# --- Prompt user for date input ---
def prompt_date(prompt_text):
    while True:
        date_str = input(f"{prompt_text} (YYYYMMDD): ")
        try:
            return datetime.strptime(date_str, "%Y%m%d")
        except ValueError:
            print("❌ Invalid format. Please use YYYYMMDD.")

# --- Clean MRNs ---
def clean__df(df):
    df["MRN"] = df["MRN"].astype(str).str.lstrip("0")
    return df

# --- Run SQL query, return cleaned DataFrame ---
def run_query_and_return(query):
    conn_str = (
        "mssql+pyodbc://@SBNC-sql/NGProd?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
    )
    try:
        engine = create_engine(conn_str)
        df = pd.read_sql_query(query, con=engine)
        clean__df(df)
        return df
    except Exception as e:
        print("❌ Failed to run SQL:", e)
        return pd.DataFrame()

# --- Main Execution Logic ---
def run_main_template_query():
    lower_date = prompt_date("Enter the Date of the Mobile Dental Event")
    upper_date = datetime.today() - timedelta(days=182)

    lower_str_sql = lower_date.strftime("%Y%m%d")
    upper_str_sql = upper_date.strftime("%Y%m%d")
    today_str = datetime.today().strftime('%Y-%m-%d')

    file_date_range1 = lower_date.strftime('%Y-%m-%d')
    file_date_range2 = f"{upper_date.strftime('%Y-%m-%d')}_to_{today_str}"

    output_file_dental = fr'C:\Reports\Dental Booking Analysis\Dental Booking Analysis {file_date_range1}.xlsx'
    output_file_medical = fr'C:\Reports\Dental Booking Analysis\Kept Medical Appointments {file_date_range2}.xlsx'
    comparison_output = fr'C:\Reports\Dental Booking Analysis\MRN_Comparison_{file_date_range1}.xlsx'

    # --- Queries ---
    sql_query_dental = f"""
    SELECT 
        x.description AS [Provider Name],
        m.event AS [Appointment Name], 
        l.location_name AS [Location Name], 
        z.appt_date AS [Appointment Date], 
        z.begintime, 
        z.appt_kept_ind AS [Kept Status?],
        z.description AS [Full Patient Name],
        q.date_of_birth,
        z.workflow_status, 
        pp.med_rec_nbr AS [MRN], 
        z.cancel_ind, 
        z.delete_ind
    FROM appointments z
    INNER JOIN location_mstr l ON l.location_id = z.location_id
    INNER JOIN provider_mstr x ON x.provider_id = z.rendering_provider_id
    INNER JOIN events m ON m.event_id = z.event_id
    INNER JOIN person q ON q.person_id = z.person_id
    FULL JOIN patient pp ON pp.person_id = q.person_id
    FULL JOIN patient_encounter pe ON pe.person_id = z.appt_id
    WHERE 
        z.appt_date = '{lower_str_sql}' AND 
        l.location_name LIKE '%Dental%' AND 
        z.cancel_ind = 'N' 
    ORDER BY z.appt_date ASC;
    """

    sql_query_medical_last_12mo = f"""
    SELECT DISTINCT pp.med_rec_nbr AS [MRN]
    FROM appointments z
    INNER JOIN location_mstr l ON l.location_id = z.location_id
    FULL JOIN patient pp ON pp.person_id = z.person_id
    WHERE 
        z.appt_date > '{upper_str_sql}' AND 
        l.location_name NOT LIKE '%Dental%' AND 
        z.appt_kept_ind = 'Y'
    """

    sql_query_medical_all_time = f"""
    SELECT DISTINCT pp.med_rec_nbr AS [MRN]
    FROM appointments z
    INNER JOIN location_mstr l ON l.location_id = z.location_id
    FULL JOIN patient pp ON pp.person_id = z.person_id
    WHERE 
        l.location_name NOT LIKE '%Dental%' AND 
        z.appt_kept_ind = 'Y'
    """

    # --- Run Queries ---
    df_dental_raw = run_query_and_return(sql_query_dental)
    print("✅ Retrieved: Dental Schedule")

    df_medical_12mo = run_query_and_return(sql_query_medical_last_12mo)
    df_medical_all = run_query_and_return(sql_query_medical_all_time)

    # --- Clean and deduplicate MRNs ---
    dental_mrns = set(df_dental_raw['MRN'].dropna())
    medical_mrns_12mo = set(df_medical_12mo['MRN'].dropna())
    medical_mrns_all = set(df_medical_all['MRN'].dropna())

    seen_in_both = dental_mrns & medical_mrns_all
    seen_in_both_not_in_last_12mo = seen_in_both - medical_mrns_12mo

    # --- Filter Data ---
    seen_in_both_df = df_dental_raw[df_dental_raw["MRN"].isin(seen_in_both)].copy()
    not_seen_med_12mo_df = df_dental_raw[df_dental_raw["MRN"].isin(seen_in_both_not_in_last_12mo)].copy()

    # --- Export final MRN comparison workbook ---
    export_to_excel_simple(
        {
            "Seen in Both": seen_in_both_df,
            "Seen in Dental, Not Medical 6mo": not_seen_med_12mo_df
        },
        comparison_output
    )
    print(f"[✓] Comparison workbook exported: {comparison_output}")

# --- Run ---
if __name__ == "__main__":
    run_main_template_query()